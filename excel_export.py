"""
Excel Export Module for Bay & Basin Rosters
Generates formatted Excel files matching the current roster format
"""

from datetime import datetime, timedelta
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from roster_assignment import RosterAssignment, StaffMember


def export_roster_to_excel(roster: RosterAssignment, filename: str = None) -> str:
    """
    Export roster to Excel file in Bay & Basin format
    
    Args:
        roster: RosterAssignment object
        filename: Optional filename (generates one if not provided)
    
    Returns:
        Path to created Excel file
    """
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Bay_Basin_Roster_{timestamp}.xlsx"
    
    try:
        wb = Workbook()
        
        # Create sheets
        create_roster_sheet(wb, roster)
        create_coverage_sheet(wb, roster)
        create_summary_sheet(wb, roster)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Save
        wb.save(filename)
        return filename
    except Exception as e:
        print(f"Error creating Excel file: {e}")
        import traceback
        traceback.print_exc()
        raise


def create_roster_sheet(wb: Workbook, roster: RosterAssignment):
    """Create the main roster view sheet"""
    ws = wb.active
    ws.title = "Roster"
    
    # Title
    ws['A1'] = "Bay & Basin Roster"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Period: {roster.roster_start_date.strftime('%d/%m/%Y')} - {roster.roster_end_date.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(size=12)
    
    # Calculate all dates in roster
    num_days = (roster.roster_end_date - roster.roster_start_date).days + 1
    dates = [roster.roster_start_date + timedelta(days=i) for i in range(num_days)]
    
    # Header row with dates
    row = 4
    ws.cell(row, 1, "Staff Name")
    ws.cell(row, 2, "Role")
    ws.cell(row, 3, "Current Line")
    
    for i, date in enumerate(dates):
        col = 4 + i
        ws.cell(row, col, date.strftime('%a\n%d/%m'))
        ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row, col).font = Font(bold=True, size=9)
    
    # Style header row
    for col in range(1, 4 + len(dates)):
        cell = ws.cell(row, col)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add staff schedules
    row = 5
    
    # Rotating roster staff
    for staff in roster.staff:
        if not hasattr(staff, 'is_fixed_roster') or staff.is_fixed_roster:
            continue
            
        ws.cell(row, 1, staff.name)
        ws.cell(row, 2, staff.role)
        
        if staff.assigned_line:
            ws.cell(row, 3, f"Line {staff.assigned_line}")
        else:
            ws.cell(row, 3, "Not assigned")
        
        # Get schedule - pass the staff object, not the name!
        schedule_list = roster.get_staff_schedule(staff)
        schedule = {date: shift for date, shift in schedule_list}
        
        for i, date in enumerate(dates):
            col = 4 + i
            shift = schedule.get(date, 'O')
            
            cell = ws.cell(row, col)
            
            if shift == 'D':
                cell.value = "Day"
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif shift == 'N':
                cell.value = "Night"
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif shift == 'LEAVE':
                cell.value = "Leave"
                cell.fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
            else:
                cell.value = "Off"
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
        
        row += 1
    
    # Fixed roster staff
    for staff in roster.staff:
        if not hasattr(staff, 'is_fixed_roster') or not staff.is_fixed_roster:
            continue
            
        ws.cell(row, 1, staff.name)
        ws.cell(row, 2, staff.role)
        ws.cell(row, 3, "Fixed")
        
        # Get schedule - pass the staff object, not the name!
        schedule_list = roster.get_staff_schedule(staff)
        schedule = {date: shift for date, shift in schedule_list}
        
        for i, date in enumerate(dates):
            col = 4 + i
            shift = schedule.get(date, 'O')
            
            cell = ws.cell(row, col)
            
            if shift == 'D':
                cell.value = "Day"
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif shift == 'N':
                cell.value = "Night"
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            elif shift == 'LEAVE':
                cell.value = "Leave"
                cell.fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
            else:
                cell.value = "Off"
            
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size=9)
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    for i in range(len(dates)):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 8


def create_coverage_sheet(wb: Workbook, roster: RosterAssignment):
    """Create coverage analysis sheet"""
    ws = wb.create_sheet("Coverage")
    
    ws['A1'] = "Coverage Analysis"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Minimum required: {roster.min_paramedics_per_shift} per shift"
    
    # Headers
    row = 4
    ws.cell(row, 1, "Date")
    ws.cell(row, 2, "Day")
    ws.cell(row, 3, "Day Shift")
    ws.cell(row, 4, "Night Shift")
    ws.cell(row, 5, "Status")
    
    for col in range(1, 6):
        cell = ws.cell(row, col)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center')
    
    # Coverage data
    row = 5
    num_days = (roster.roster_end_date - roster.roster_start_date).days + 1
    
    for i in range(num_days):
        date = roster.roster_start_date + timedelta(days=i)
        coverage = roster.get_coverage_for_date(date)
        
        ws.cell(row, 1, date.strftime('%d/%m/%Y'))
        ws.cell(row, 2, date.strftime('%A'))
        ws.cell(row, 3, coverage['D'])
        ws.cell(row, 4, coverage['N'])
        
        # Status
        if coverage['D'] >= roster.min_paramedics_per_shift and coverage['N'] >= roster.min_paramedics_per_shift:
            ws.cell(row, 5, "✓ OK")
            ws.cell(row, 5).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            issues = []
            if coverage['D'] < roster.min_paramedics_per_shift:
                issues.append(f"Day short {roster.min_paramedics_per_shift - coverage['D']}")
            if coverage['N'] < roster.min_paramedics_per_shift:
                issues.append(f"Night short {roster.min_paramedics_per_shift - coverage['N']}")
            ws.cell(row, 5, ", ".join(issues))
            ws.cell(row, 5).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25


def create_summary_sheet(wb: Workbook, roster: RosterAssignment):
    """Create summary sheet with stats"""
    ws = wb.create_sheet("Summary", 0)  # Insert as first sheet
    
    ws['A1'] = "Bay & Basin Roster Summary"
    ws['A1'].font = Font(size=16, bold=True)
    
    row = 3
    
    # Roster period
    ws.cell(row, 1, "Roster Period:")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, f"{roster.roster_start_date.strftime('%d/%m/%Y')} - {roster.roster_end_date.strftime('%d/%m/%Y')}")
    row += 1
    
    num_days = (roster.roster_end_date - roster.roster_start_date).days + 1
    ws.cell(row, 1, "Duration:")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, f"{num_days} days ({num_days // 7} weeks)")
    row += 2
    
    # Staff counts
    ws.cell(row, 1, "Staff Summary")
    ws.cell(row, 1).font = Font(size=14, bold=True)
    row += 1
    
    rotating = [s for s in roster.staff if hasattr(s, 'is_fixed_roster') and not s.is_fixed_roster]
    fixed = [s for s in roster.staff if hasattr(s, 'is_fixed_roster') and s.is_fixed_roster]
    
    ws.cell(row, 1, "Total Staff:")
    ws.cell(row, 2, len(roster.staff))
    row += 1
    
    ws.cell(row, 1, "Rotating Roster:")
    ws.cell(row, 2, len(rotating))
    row += 1
    
    ws.cell(row, 1, "Fixed/Casual:")
    ws.cell(row, 2, len(fixed))
    row += 2
    
    # Line assignments
    ws.cell(row, 1, "Line Assignments")
    ws.cell(row, 1).font = Font(size=14, bold=True)
    row += 1
    
    for line_num in range(1, 10):
        staff_on_line = [s for s in rotating if s.assigned_line == line_num]
        if staff_on_line:
            ws.cell(row, 1, f"Line {line_num}:")
            ws.cell(row, 2, f"{len(staff_on_line)} staff")
            ws.cell(row, 3, ", ".join([s.name for s in staff_on_line]))
            row += 1
    
    row += 1
    
    # Coverage issues
    ws.cell(row, 1, "Coverage Status")
    ws.cell(row, 1).font = Font(size=14, bold=True)
    row += 1
    
    issues = roster.check_coverage()
    if not issues:
        ws.cell(row, 1, "✓ All shifts adequately covered")
        ws.cell(row, 1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        ws.cell(row, 1, f"⚠ {len(issues)} coverage issue(s) - see Coverage sheet")
        ws.cell(row, 1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40


# Quick test
if __name__ == "__main__":
    print("Excel export module loaded successfully")
    print("Use: export_roster_to_excel(roster) to generate Excel file")
